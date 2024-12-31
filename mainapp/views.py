import os
import requests
from functools import cache
import json
from django.http import JsonResponse
from django.shortcuts import render
from rest_framework.response import Response
from rest_framework.authtoken.models import Token
from rest_framework.views import APIView
from rest_framework.authtoken.views import ObtainAuthToken
from rest_framework import status
from mainapp.serializers import RoleSerializer,SystemUserSerializer
from rest_framework.decorators import api_view,permission_classes
from django.contrib.auth.models import User
from django.utils.crypto import get_random_string
from django.core.mail import send_mail
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth import get_user_model
from datetime import timedelta
from django.utils import timezone
from pos import settings
from .models import SystemUser,Roles,Team,TeamJoin,Member,Product,Category,Department,Employee,Customer,PlaceOrder,Blog
from django.contrib.auth.hashers import make_password
from django.template.loader import render_to_string
from django.core.mail import EmailMessage
from django.views import View
from django.forms.models import model_to_dict
from django.shortcuts import get_object_or_404
from rest_framework import viewsets
from mainapp.serializers import ProductSerializer,CategorySerializer,DepartmentSerializer,EmployeeSerializer,BlogSerializer
from rest_framework.decorators import action
from rest_framework.viewsets import ModelViewSet
from django_filters.rest_framework import DjangoFilterBackend
from .filters import ProductFilter
from rest_framework.filters import SearchFilter,OrderingFilter
from rest_framework.pagination import PageNumberPagination
import threading
from .models import ParentChildCategory,ParentChildProduct
from django.shortcuts import HttpResponse
from openpyxl import Workbook
import openpyxl
from rest_framework.permissions import AllowAny
from openpyxl.styles import Font, Alignment
from django.utils.dateparse import parse_datetime
from rest_framework.parsers import MultiPartParser, FormParser
import pytz
from datetime import datetime
from dateutil import parser
from django.utils.timezone import make_aware


def Home(request):
    return render(request,'home.html')

class DynamicPageNumberPagination(PageNumberPagination):
    page_size_query_param = 'page_size'

    def get_paginated_response(self, data):
        return Response({
            'current_page': self.page.number,
            'total_pages': self.page.paginator.num_pages,
            'total_items': self.page.paginator.count,
            # 'page_size': self.page_size,
            'results': data,
        })

class DepartmentView(ModelViewSet):
    queryset = Department.objects.all()
    serializer_class = DepartmentSerializer

class EmployeeView(ModelViewSet):
    queryset = Employee.objects.filter(is_active=True)
    serializer_class = EmployeeSerializer

class LoginView(ObtainAuthToken):
    def post(self, request, *args, **kwargs):
        # Validate the data and check user credentials
        serializer = self.serializer_class(data=request.data, context={'request': request})
        serializer.is_valid(raise_exception=True)
        user = serializer.validated_data['user']
        
        # Check if user is a superuser
        if user.is_superuser:
            # Handle logic if the user is a superuser, like logging them in differently if needed
            # For now, returning the token and user info as usual
            token, created = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user_id': user.id,
                'email': user.email,
                'is_superuser': user.is_superuser  # Add superuser status
            })
        else:
            # Regular user login logic
            token, created = Token.objects.get_or_create(user=user)
            return Response({
                'token': token.key,
                'user_id': user.id,
                'email': user.email
            })

class RoleCreateView(APIView):
    def post(self, request):
        serializer = RoleSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
class SystemUserCreateView(APIView):
    def post(self, request):
        serializer = SystemUserSerializer(data=request.data)
        if serializer.is_valid():
            user = serializer.save()
            return Response({
                "id": user.id,
                "username": user.username,
                "email": user.email,
                "mobile_no": user.mobile_no,
                "role": user.role_id.name
            }, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

@api_view(['POST'])
def create_system_user(request):
    if request.method == 'POST':
        username = request.data.get('username')
        email = request.data.get('email')
        mobile_no = request.data.get('mobile_no')
        profile_image = request.FILES.get('profile_image')
        role_id = request.data.get('role_id')
        password = request.data.get('password')

        if not all([username, email, mobile_no, role_id, password]):
            return Response({"error": "All fields are required."}, status=status.HTTP_400_BAD_REQUEST)

        if SystemUser.objects.filter(email=email).exists() or SystemUser.objects.filter(mobile_no=mobile_no).exists():
            return Response({"error": "Email or mobile number already exists."}, status=status.HTTP_400_BAD_REQUEST)

        try:
            role = Roles.objects.get(id=role_id)
        except Roles.DoesNotExist:
            return Response({"error": "Invalid role ID."}, status=status.HTTP_400_BAD_REQUEST)

        user = SystemUser.objects.create(
            username=username,
            email=email,
            mobile_no=mobile_no,
            profile_image=profile_image,
            role_id=role,
            password=make_password(password)
        )

        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
            "role": user.role_id.name
        }, status=status.HTTP_201_CREATED)

@api_view(['GET'])
def get_user(request, user_id):
    try:
        user = SystemUser.objects.get(id=user_id)
        if user.is_deleted:
            return Response({"message":"User not found"})
        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
            "role": user.role_id.name
        })
    except SystemUser.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

@api_view(['PUT'])
def update_user(request, user_id):
    try:
        user = SystemUser.objects.get(id=user_id)
        if user.is_deleted:
            return Response({"message":"User not found"})
        user.username = request.data.get('username', user.username)
        user.email = request.data.get('email', user.email)
        user.mobile_no = request.data.get('mobile_no', user.mobile_no)
        user.profile_image = request.FILES.get('profile_image', user.profile_image)
        
        password = request.data.get('password')
        if password:
            user.password = make_password(password)
        
        user.save()

        return Response({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
            "role": user.role_id.name
        })
    except SystemUser.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)

@api_view(['DELETE'])
def delete_user(request, user_id):
    try:
        user = SystemUser.objects.get(id=user_id)
        if user.is_deleted:
            return Response({"message": "User is already deleted."}, status=status.HTTP_400_BAD_REQUEST)
        user.delete()
        return Response({"message": "User soft-deleted successfully."}, status=status.HTTP_200_OK)
    except SystemUser.DoesNotExist:
        return Response({"error": "User not found."}, status=status.HTTP_404_NOT_FOUND)
    
@api_view(['GET'])
def fetch_user_roles(request):
    user_data = SystemUser.objects.filter(is_deleted=False)
    users_list = []
    for user in user_data:
        users_list.append({
            "id": user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
            "role": user.role_id.name if user.role_id else None,
        })
    return Response(users_list)

@api_view(['GET'])
def role_specific_fetch_users_byname(request,name):
    role_data = Roles.objects.get(name=name)
    user_data = SystemUser.objects.filter(role_id=role_data.id,is_deleted=False)
    user_list =[]
    for user in user_data:
        user_list.append({
            "id":user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
        })
    return Response(user_list)

@api_view(['GET'])
def role_specific_fetch_users_byid(request,id):
    user_data = SystemUser.objects.filter(role_id=id,is_deleted=False)
    user_list =[]
    for user in user_data:
        user_list.append({
            "id":user.id,
            "username": user.username,
            "email": user.email,
            "mobile_no": user.mobile_no,
            "profile_image": user.profile_image.url if user.profile_image else None,
        })
    return Response(user_list)  

def email_thread(user,otp,email):
    try:
            expiry_time = timezone.now() + timedelta(minutes=5)
            user.otp = otp
            user.expiry_time = expiry_time
            user.save()

            pdf_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'blankpdf.pdf')

            if not os.path.exists(pdf_path):
                return JsonResponse({'error': 'Static PDF file not found'}, status=500)

            if not os.path.exists(pdf_path):
                return JsonResponse({'error': 'Static PDF file not found'}, status=500)

            with open(pdf_path, 'rb') as pdf_file:
                attachment_content = pdf_file.read()
                attachment_filename = "blankpdf.pdf"

            if user.is_active:
                html_content = render_to_string('otp_email_template.html', {
                    'otp': otp,
                    'expiry_time': '5 minutes',
                    'username': user.username,  
                })
                email_message = EmailMessage(
                    subject='Your OTP Code',
                    body=html_content,
                    from_email='your-email@gmail.com',
                    to=[email],
                )
                email_message.content_subtype = 'html'
            else:
                html_content = render_to_string('non_active_email_template.html', {
                    'username': user.username,
                    'activation_link': f'http://yourwebsite.com/activate/'  # Replace with appropriate activation URL
                })
                email_message = EmailMessage(
                    subject='Action Required: Activate Your Account',
                    body=html_content,
                    from_email='your-email@gmail.com',
                    to=[email],
                )
                email_message.content_subtype = 'html'

            email_message.attach(attachment_filename, attachment_content, "application/pdf")

            email_message.send()
    except Exception as e:
        JsonResponse({"error":"Error is found when sending mail"})

@api_view((['POST']))
def send_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')

            if not email:
                return JsonResponse({'error': 'Email is required'}, status=400)

            user = SystemUser.objects.filter(email=email).first()
            if not user:
                user=User.objects.filter(email=email).first()
                if not user:
                    return JsonResponse({'error': 'User not found'}, status=404)

            otp = get_random_string(length=6, allowed_chars='0123456789')

            threading.Thread(target=email_thread,args=(user,otp,email)).start()

            return JsonResponse({'message': 'Email sent successfully with attachment'}, status=200)

        except Exception as e:
            return JsonResponse({'error': f'Failed to process request: {str(e)}'}, status=500)

    return JsonResponse({'error': 'Invalid request method'}, status=405)

@csrf_exempt
def verify_otp(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            email = data.get('email')
            otp = data.get('otp')

            if not email or not otp:
                return JsonResponse({'error': 'Email and OTP are required'}, status=400)

            user = SystemUser.objects.filter(email=email).first()
            if not user:
                return JsonResponse({'error': 'User not found'}, status=404)

            if user.otp != int(otp):
                return JsonResponse({'error': 'Invalid OTP'}, status=400)

            if timezone.now() > user.expiry_time:
                return JsonResponse({'error': 'OTP has expired'}, status=400)

            new_otp = get_random_string(length=6, allowed_chars='0123456789')
            expiry_time = timezone.now() + timedelta(minutes=5)

            user.otp = new_otp
            user.expiry_time = expiry_time
            user.save()

            return JsonResponse({'message': 'OTP verified successfully. New OTP generated.'}, status=200)

        except Exception as e:
            return JsonResponse({'error': f'Failed to process request: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Invalid request method'}, status=405)

class RolesCRUD(APIView):
    def get(self,request,id=None):
        if id:
            # role = get_object_or_404(Roles,id=id)
            # role = Roles.objects.get(id=id)
            role = Roles.objects.filter(id=id, is_deleted=False).get()
            return JsonResponse(model_to_dict(role))
        # rolealldata = Roles.objects.all()
        rolealldata = Roles.objects.filter(is_deleted=False)
        role_list = list(rolealldata.values())
        return JsonResponse(role_list,safe=False)
    
    def post(self,request):
        try:
            data = json.loads(request.body)
            name = data.get('name')
            if not name:
                return Response({"message":"Name is required provide it"})
            data = Roles.objects.create(name=name)
            return JsonResponse(model_to_dict(data),status=201)
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
    
    def put(self,request,id):
        roledata = get_object_or_404(Roles,id=id)
        try:
            if roledata is None:
                return Response("Your are not providing the id")
            data = json.loads(request.body)
            name = data.get("name")
            isactive = data.get("is_active")
            if name is not None:
                roledata.name=name
            if isactive is not None:
                roledata.is_active=isactive
            roledata.save()
            return JsonResponse(model_to_dict(roledata))
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        
    def delete(self,request,id):
        role_data = get_object_or_404(Roles,id=id)
        role_data.delete()
        return Response("Role is deleted successfully")    
    

@api_view(['POST'])
def update_team_members(request):
    team_name = request.data.get("team")
    members = request.data.get("members")

    if team_name is None or members is None:
        return Response({"Error":"Provide team and team members "})
    
    # team = Team.objects.get_or_create(name=team_name)
    team, created = Team.objects.get_or_create(name=team_name)


    old_team_members = TeamJoin.objects.filter(team=team)
    members_dict = {}

    for team_member in old_team_members:
        member_name = team_member.member.name
        members_dict[member_name] = team_member

    incoming_member=[]
    for member_data in members:
        name = member_data.get("name")
        points = member_data.get("points") 

        if name is None or points is None:
            continue

        incoming_member.append(name)

        if name in members_dict:
            team_join = members_dict[name]
            if team_join.points != points:
                team_join.points = points
                team_join.save()
        else:
            # member = Member.objects.get_or_create(name=name)
            member, _ = Member.objects.get_or_create(name=name)
    
            TeamJoin.objects.create(team=team,member=member,points=points)
    
    for member_name, team_join in members_dict.items():
        if member_name not in incoming_member:
            team_join.delete()

    return Response({"message": "Team members updated successfully."})

@api_view(['GET'])
def get_all_teams_and_members(request):
    teams = Team.objects.all()
    result = []
    for team in teams:
        team_data = {"team": team.name, "members": []}
        team_joins = TeamJoin.objects.filter(team=team).select_related('member')
        for join in team_joins:
            team_data["members"].append({
                "name": join.member.name,
                "points": join.points
            })
        result.append(team_data)
    return Response(result)

class ProductViewset(viewsets.ViewSet):
    def list(self, request):
        data = request.GET.get('sort',None)
        products = Product.objects.all()

        name = request.GET.get('name', None)
        price = request.GET.get('price', None)
        is_active = request.GET.get('is_active', None)
        
        if name is not None:
            products = products.filter(prod_name__icontains = name)

        if price is not None:
            try:
                price = int(price)
                products = products.filter(price = price)
            except ValueError:
                return Response({"error": "Invalid price format"}, status=400)

        if is_active is not None:
            try:
                is_active = bool(int(is_active))
                products = products.filter(is_active=is_active)
            except ValueError:
                return Response({"error": "is_active must be 0 or 1"}, status=400)

        if data == "desc":
            products = products.order_by('-price')
        else:
            products = products.order_by('price')
        
        # serializer = ProductSerializer(products, many=True)
        # return Response(serializer.data)

        paginator = DynamicPageNumberPagination()
        paginated_products = paginator.paginate_queryset(products, request)

        serializer = ProductSerializer(paginated_products, many=True)
        return paginator.get_paginated_response(serializer.data)
    
    def create(self, request):
        serializer = ProductSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    
    def retrieve(self, request, pk=None):
        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({"error": "Product not found"}, status=status.HTTP_404_NOT_FOUND)
        serializer = ProductSerializer(product)
        return Response(serializer.data)

    def destroy(self , request , pk=None):
        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({"error":"Product not found"}, status=status.HTTP_404_NOT_FOUND)
        product.delete()
        return Response({"Message":"Your data is successfully deleted"})
    
    def update(self , request , pk=None):
        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({"error":"Product not found"},status=status.HTTP_404_NOT_FOUND)
        serializer = ProductSerializer(product,data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
    
    def partial_update(self , request , pk=None):
        try:
            product = Product.objects.get(pk=pk)
        except Product.DoesNotExist:
            return Response({"error":"Product not found"},status=status.HTTP_404_NOT_FOUND)
        serializer = ProductSerializer(product,data=request.data,partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)

class CategoryViewset(viewsets.ViewSet):
    def list(self,request):
        category = Category.objects.all()
        name = request.query_params.get('name', None)
        is_active = request.query_params.get('is_active', None)

        if name is not None:
            categories = categories.filter(cat_name__icontains=name)  # Search by category name (case-insensitive)
        
        if is_active is not None:
            categories = categories.filter(is_active=is_active)  # Filter by is_active status

        serializer = CategorySerializer(category,many=True)
        return Response(serializer.data)
    
    def create(self,request):
        serializer = CategorySerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
    
    def update(self,request,pk=None):
        try:
            category = Category.objects.get(id=pk)
        except Category.DoesNotExist:
            return Response({"error":"Category not found"})
        serializer = CategorySerializer(category,data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
    
    def destroy(self,request,pk=None):
        try:
            category = Category.objects.get(id=pk)
        except Category.DoesNotExist:
            return Response({"error":"Category not found"})
        category.delete()
        return Response("Your data deleted successfully")
    
    def partial_update(self,request,pk=None):
        try:
            category = Category.objects.get(id=pk)
        except Category.DoesNotExist:
            return Response({"error":"Category not found"})
        serializer = CategorySerializer(category,data=request.data,partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response(serializer.errors)
    
class ProductModelViewset(ModelViewSet):
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    filter_backends = [DjangoFilterBackend,SearchFilter,OrderingFilter]
    filterset_class = ProductFilter
    pagination_class = DynamicPageNumberPagination

    search_fields = ['name']
    ordering_fields = ['price', 'name']

@api_view(['GET'])
def ProductPaginationData(request):
    products = Product.objects.all()
    paginator = DynamicPageNumberPagination()

    paginated_products = paginator.paginate_queryset(products, request)

    serializer = ProductSerializer(paginated_products, many=True)
    return paginator.get_paginated_response(serializer.data)

class PlacedOrderView(APIView):
    def post(self, request):
        try:
            data = request.data
            order_status = data.get('order_status', 'PENDING')
            customer_id = data.get('cust_id')
            product_id = data.get('prod_id')
            quantity = data.get('quantity')
            discount = data.get('discount', 0)
            payment_type = data.get('payment_type')

            if not customer_id or not product_id or not quantity or not payment_type:
                return JsonResponse({"error": "Missing required fields"}, status=400)

            try:
                customer = Customer.objects.get(id=customer_id)
            except Customer.DoesNotExist:
                return JsonResponse({'error': 'Customer not found'}, status=404)

            try:
                product = Product.objects.get(id=product_id)
            except Product.DoesNotExist:
                return JsonResponse({'error': 'Product not found'}, status=404)

            # Calculate the order details
            amount = product.price
            delivery_charges = 50 if payment_type in ['ONLINE', 'COD'] else 0
            discount_amount = (discount * (amount * quantity)) / 100
            total_amount = (amount * quantity) + delivery_charges - discount_amount

            # Create the order
            placeorder=PlaceOrder.objects.create(
                order_status=order_status,
                cust_id=customer,
                prod_id=product,
                quantity=quantity,
                amount=amount,
                total_amount=total_amount,
                delivery_charges=delivery_charges,
                discount=discount,
                payment_type=payment_type
            )
            return Response({"message": "Your order is confirmed!"}, status=201)
        except Exception as e:
            return Response({"error": f"An error occurred: {str(e)}"}, status=500)
        
class OrderHistory(APIView):
    def get(self,request):
        order = PlaceOrder.objects.all()
        i=1
        print(order)
        allorder = []
        for single_order in order:
            cust_detail = single_order.cust_id
            prod_detail = single_order.prod_id

            allorder.append({f'Order {i}':{
                "customer_name": cust_detail.cust_name,
                "customer_address": cust_detail.cust_address,
                "product_name": prod_detail.prod_name,
                "amount": single_order.amount,
                "quantity": single_order.quantity,
                "total_amount": single_order.total_amount,
                "payment_type": single_order.payment_type,
                "payment_date": single_order.payment_date,
            }})
            i=i+1
        return Response(allorder)
    
class BlogViewSet(ModelViewSet):
    queryset = Blog.objects.all()
    serializer_class = BlogSerializer
    
    def get_queryset(self):
        slug = self.request.GET.get('slug', None)
        if slug:
            return Blog.objects.filter(slug=slug)
        else:
            return Blog.objects.all()

class DisplayProductByCategory(APIView):
    def get(self, request, id=None):
        if not id:
            return JsonResponse({"error": "You did not provide a valid ID."}, status=400)
        try:
            parent_category = ParentChildCategory.objects.get(id=id)
            print(f'parent category = {parent_category}')
            subcategories = ParentChildCategory.objects.filter(parent=parent_category)
            print(f'subcategories ={list(subcategories)}')
            categories = [parent_category] + list(subcategories)
            print(f'categories ={categories}')
            products = ParentChildProduct.objects.filter(cat__in=categories).values('id', 'name', 'price', 'cat__name')

            return JsonResponse({
                "parent_category": parent_category.name,
                "products": list(products)
            })
        except ParentChildCategory.DoesNotExist:
            return JsonResponse({"error": "Invalid category ID provided."}, status=404)
        
class OrderExcel(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'PlaceOrders'

        headers = [
            'Order ID', 'Customer Name', 'Customer Address', 'Product Name', 'Order Status',
            'Amount', 'Quantity', 'Total Amount', 'Delivery Charges', 'Discount', 'Payment Type', 'Payment Date'
        ]
        sheet.append(headers)

        orders = PlaceOrder.objects.all()

        status_counts = {'Pending': 0,'PENDING':0, 'COMPLETED': 0, 'CANCELED': 0}
        total_orders = 0

        for order in orders:
            total_orders += 1

            # Update order status counts
            if order.order_status in status_counts:
                status_counts[order.order_status] += 1

            # Format payment date
            formatted_date = order.payment_date.strftime('%A, %d %B %Y, %I:%M%p')

            sheet.append([
                order.id,
                order.cust_id.cust_name if order.cust_id else 'N/A',
                order.cust_id.cust_address if order.cust_id else 'N/A',
                order.prod_id.prod_name if order.prod_id else 'N/A',
                order.order_status,
                order.amount,
                order.quantity,
                order.total_amount,
                order.delivery_charges,
                order.discount,
                order.payment_type,
                formatted_date
            ])

        # Add footer information (total orders and order status counts)
        sheet.append([])
        sheet.append([f'Total Orders = {total_orders}'])
        sheet.append([f'Pending Orders = {status_counts["Pending"] + status_counts["PENDING"]}'])
        sheet.append([f'Completed Orders = {status_counts["COMPLETED"]}'])
        sheet.append([f'Canceled Orders = {status_counts["CANCELED"]}'])

        # Make header row bold
        for cell in sheet["1:1"]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Prepare the response to send the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="placeorders.xlsx"'

        workbook.save(response)
        return response
 
class ProductExcel(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Products'

        # Define headers
        headers = ['Product ID', 'Product Name', 'Price', 'Description', 'Is Active']
        sheet.append(headers)

        products = Product.objects.all()

        # Loop through the products and append each one to the sheet
        for product in products:
            sheet.append([
                product.id,
                product.prod_name,
                product.price,
                product.desc,
                'Yes' if product.is_active else 'No'
            ])

        # Bold the header row
        for cell in sheet["1:1"]:
            cell.font = openpyxl.styles.Font(bold=True)

        # Prepare the response to send the Excel file
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="products.xlsx"'

        workbook.save(response)
        return response

class UploadProductExcel(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        # Load the Excel file
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active

        # Loop through rows starting from row 2 (after the header)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            print("Processing row:", row)  # Debug: print each row

            # Extract values from the row
            product_name = row[1]  # Product Name
            price = row[2]  # Price
            description = row[3]  # Description
            is_active = True if row[4].lower() == 'yes' else False  # Is Active

            # Create a new Product record
            Product.objects.create(
                prod_name=product_name,
                price=price,
                desc=description,
                is_active=is_active
            )

        return JsonResponse({"message": "Products successfully uploaded"}, status=200)
    
class UploadOrderExcel(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        if not excel_file:
            return JsonResponse({"error": "No file uploaded"}, status=400)

        try:
            # Load the Excel workbook
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active

            for row in sheet.iter_rows(min_row=2, values_only=True):
                customer_name = row[1]
                customer_address = row[2]
                product_name = row[3]
                order_status = row[4].upper() if row[4] else 'PENDING'
                amount = row[5]
                quantity = row[6]
                total_amount = row[7]
                delivery_charges = row[8]
                discount = row[9]
                payment_type = row[10].upper() if row[10] else 'CASH'
                payment_date_str = row[11]  # Keep raw payment date as string for debugging

                try:
                    # Explicit date parsing
                    payment_date = parser.parse(payment_date_str)  # Automatically handles most formats
                    payment_date = make_aware(payment_date)  # Make timezone-aware if settings.USE_TZ=True
                except Exception as e:
                    return JsonResponse({"error": f"Invalid date format: {str(e)}"}, status=400)

                customer = Customer.objects.filter(cust_name=customer_name, cust_address=customer_address).first()
                product = Product.objects.filter(prod_name=product_name).first()

                if customer and product:
                    PlaceOrder.objects.create(
                        order_status=order_status,
                        payment_type=payment_type,
                        cust_id=customer,
                        prod_id=product,
                        quantity=quantity,
                        amount=amount,
                        total_amount=total_amount,
                        delivery_charges=delivery_charges,
                        discount=discount,
                        payment_date=payment_date
                    )

            return JsonResponse({"message": "Orders successfully uploaded"}, status=200)

        except Exception as e:
            return JsonResponse({"error": f"Error processing file: {str(e)}"}, status=500)
